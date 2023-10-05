import requests
import json
import hashlib
import base64
import pytest
import pandas as pd
from conftest import contractors_inn
import openpyxl

def test_contractors_create(api_credentials):
    log = api_credentials['LOG']
    password = api_credentials['PASS']
    base_url = api_credentials['BASE_URL']
    org_id = api_credentials['ORG_ID']

    # Аутентификация и получение токена
    response_auth = requests.post(base_url + "/oauth/token", data={
        "username": log,
        "password": base64.b64encode(hashlib.sha256(password.encode()).digest()).decode(),
        "client_id": "cashdesk-rest-client",
        "client_secret": "cashdesk-rest-client",
        "grant_type": "password"
    })
    assert response_auth.status_code == 200

    data = json.loads(response_auth.text)
    assert "access_token" in data
    token = data["access_token"]
    print(f"Access token: {token}")

    data_to_send = []
    headers = {
        "Authorization": "Bearer " + token,
    }

    file_path = "ЗАКУПКИ.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['поставщики']

    # Перебираем уникальные наименования поставщиков и отправляем запросы на создание поставщиков
    for contractor_id, inn in contractors_inn().items():
        # Получаем уникальные наименования из столбца с id (нумерация с 1)
        unique_supplier_names = set(sheet.cell(row=row, column=1).value for row in range(3, sheet.max_row + 1) if
                                    sheet.cell(row=row, column=1).value)

        # Заменяем None (пустые ячейки) на пустую строку
        unique_supplier_names = [name if name is not None else '' for name in unique_supplier_names]

        supplier_ids = [sheet.cell(row=row, column=3).value for row in range(3, sheet.max_row + 1)]
        supplier_ids = [int(id) if str(id).isdigit() else 0 for id in supplier_ids]

        # Получаем соответствующие инн из структуры inn_contractors
        inn = contractors_inn().get(contractor_id, '')
        print("Unique Supplier Names:", unique_supplier_names)
        print("Supplier IDs:", supplier_ids)
        for supplier_name, supplier_id in zip(unique_supplier_names, supplier_ids):

            if supplier_id == 0 or int(supplier_id) != int(contractor_id):
                continue

            url = f'{base_url}/protected/contractors'
            payload = {
                "item": {
                    "fullName": supplier_name,
                    "groupTypes": ["SUPPLIER"],
                    "inn": inn,
                    "organizationId": org_id,
                    "shortName": supplier_name,
                    "visible": True
                },
                "organizationId": org_id
            }

            print(f'Sending request to: {url}')
            data_to_send.append(payload)

            response = requests.post(url, json=payload, headers=headers)

            if response.status_code == 200:
                response_json = response.json()
                supplier_id = response_json.get('id')

                # Находим индекс строки, соответствующей текущему поставщику
                row_index = unique_supplier_names.index(
                    supplier_name) + 3


                sheet.cell(row=row_index, column=3, value=supplier_id)

                print(f'Created item for supplier: {supplier_name}, ID: {supplier_id}')

    # Сохраняем изменения в файл
    wb.save(file_path)